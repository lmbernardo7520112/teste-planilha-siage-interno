import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Cria um DataFrame vazio com a estrutura desejada
colunas = [
    "Nome",
    "Matemática B1", "Matemática B2", "Média Matemática",
    "Português B1", "Português B2", "Média Português",
    "História B1", "História B2", "Média História"
]

# Cria um DataFrame vazio com as colunas definidas
df = pd.DataFrame(columns=colunas)

# Cria um arquivo Excel usando openpyxl
wb = Workbook()
ws = wb.active

# Adiciona o cabeçalho ao arquivo Excel
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Adiciona fórmulas para calcular as médias
for row in range(2, 100):  # Assume até 100 linhas de alunos
    # Fórmula para Média Matemática
    ws[f'D{row}'] = f'=AVERAGE(B{row},C{row})'
    # Fórmula para Média Português
    ws[f'G{row}'] = f'=AVERAGE(E{row},F{row})'
    # Fórmula para Média História
    ws[f'J{row}'] = f'=AVERAGE(H{row},I{row})'

# Define o caminho onde o arquivo será salvo
caminho_padrao = "/mnt/c/Users/lmbernardo/Downloads"  # Caminho no WSL2 para a pasta Downloads do Windows
nome_arquivo = "planilha_notas_com_formulas.xlsx"
caminho_completo = os.path.join(caminho_padrao, nome_arquivo)

# Salva o arquivo Excel
wb.save(caminho_completo)

print(f"Arquivo Excel salvo em: {caminho_completo}")