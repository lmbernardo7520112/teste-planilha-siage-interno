import requests
import openpyxl
from app.utils.excel_utils import criar_dashboard_sec_aprovacao
from app.core.config import DISCIPLINAS, CAMINHO_PADRAO, NOME_ARQUIVO_PADRAO
import json
from pathlib import Path
import os

# Configurações
URL_ENDPOINT = "http://localhost:8000/api/v1/gerar-planilha"
CAMINHO_PLANILHA_BASE = f"{CAMINHO_PADRAO}/{NOME_ARQUIVO_PADRAO}"
CAMINHO_PLANILHA_TESTE = f"{CAMINHO_PADRAO}/planilha_teste_populada.xlsx"
CAMINHO_JSON = Path(__file__).parent / "turmas_alunos.json"

def chamar_endpoint():
    """Chama o endpoint para gerar a planilha base."""
    if os.path.exists(CAMINHO_PLANILHA_BASE):
        os.remove(CAMINHO_PLANILHA_BASE)  # Remove o arquivo existente para evitar PermissionError
    response = requests.get(URL_ENDPOINT)
    if response.status_code == 200:
        with open(CAMINHO_PLANILHA_BASE, "wb") as f:
            f.write(response.content)
        print(f"Planilha base gerada em: {CAMINHO_PLANILHA_BASE}")
    else:
        raise Exception(f"Erro ao chamar o endpoint: {response.status_code}")

def carregar_turmas():
    """Carrega as turmas do arquivo JSON."""
    with open(CAMINHO_JSON, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    return dados["turmas"]

def popular_dados_ficticios(wb, turmas):
    """Popula a planilha com dados fictícios para o B1 em todas as disciplinas."""
    for idx, turma in enumerate(turmas):
        # Linha inicial para dados da turma (após cabeçalhos e título da turma)
        linha_base = 4 + (idx * 53)  # 53 linhas por turma (1 cabeçalho + 1 título + 35 dados + 15 dashboard)
        num_alunos = len(turma["alunos"])
        
        for disciplina in DISCIPLINAS:
            ws = wb[disciplina]
            for aluno in turma["alunos"]:
                row = linha_base + int(aluno["numero"]) - 1  # Ajusta para a linha correta do aluno
                # Notas fictícias para B1 (coluna C)
                if "A" in turma["nome_turma"]:
                    ws[f"C{row}"] = 8 if int(aluno["numero"]) <= num_alunos * 0.8 else 6
                else:
                    ws[f"C{row}"] = 8 if int(aluno["numero"]) <= num_alunos * 0.2 else 6
                
                # Calcula MG (média geral) para consistência
                ws[f"H{row}"] = f"=SUM(C{row}:F{row})/4"
            
            # Taxa de aprovação (ajustado para a linha correta do dashboard)
            linha_taxa = 12 + (idx * 53)
            ws[f"O{linha_taxa}"] = f"=COUNTIF(C{linha_base}:C{linha_base + num_alunos - 1}, \">=7\")/COUNTA(C{linha_base}:C{linha_base + num_alunos - 1})"

def gerar_planilha_teste():
    """Gera a planilha de teste populada com dashboards e gráfico."""
    chamar_endpoint()
    
    wb = openpyxl.load_workbook(CAMINHO_PLANILHA_BASE)
    turmas = carregar_turmas()
    
    popular_dados_ficticios(wb, turmas)
    
    # Calcular a linha inicial do dashboard na aba SEC
    # Cada turma ocupa 40 linhas (1 título + 1 cabeçalho + 35 dados + 3 dashboard)
    # Há 7 turmas, então a última linha usada é 1 + (40 * 7) = 281
    # Começar o dashboard na linha 283 para evitar sobreposição
    LINHAS_INICIO_TABELAS = [1 + (40 * len(turmas)) + 2]
    
    ws_sec = wb["SEC"]
    criar_dashboard_sec_aprovacao(ws_sec, turmas, LINHAS_INICIO_TABELAS)
    
    wb.save(CAMINHO_PLANILHA_TESTE)
    print(f"Planilha de teste gerada em: {CAMINHO_PLANILHA_TESTE}")

if __name__ == "__main__":
    gerar_planilha_teste()