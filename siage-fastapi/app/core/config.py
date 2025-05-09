from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path

# Caminhos para arquivos e diretórios
CAMINHO_IMAGEM = Path("/home/lmbernardo/teste-planilha-siage-interno/siage-fastapi/static/images/siage_interno.png")
CAMINHO_PADRAO = "/home/lmbernardo/teste-planilha-siage-interno/siage-fastapi"
NOME_ARQUIVO_PADRAO = "planilha_notas_complexa.xlsx"

# Lista de disciplinas
DISCIPLINAS = [
    "BIO", "MAT", "FIS", "QUI", "GEO", "SOC", "HIS", 
    "FIL", "ESP", "POR", "ART", "EDF", "ING"
]

# Dicionário mapeando códigos para nomes completos das disciplinas
DISCIPLINAS_NOMES = {
    "BIO": "Biologia",
    "MAT": "Matemática",
    "FIS": "Física",
    "QUI": "Química",
    "GEO": "Geografia",
    "SOC": "Sociologia",
    "HIS": "História",
    "FIL": "Filosofia",
    "ESP": "Espanhol",
    "POR": "Português",
    "ART": "Artes",
    "EDF": "Educação Física",
    "ING": "Inglês"
}

# Colunas das abas de disciplinas
COLUNAS = [
    "Nº", "Nome do Aluno", "1º BIM", "2º BIM", "3º BIM", "4º BIM",
    "NF", "MG", "MF", "SITUAÇÃO DO ALUNO", "PF", "SF"
]

# Colunas da aba SEC
COLUNAS_SEC = [
    "Nº", "Nome do Aluno", "ATIVO", "TRANSFERIDO", "DESISTENTE", "SITUAÇÃO DO ALUNO"
]

# Configurações de largura das colunas (em centímetros)
LARGURAS_COLUNAS = {
    "Nº": 1,
    "Nome do Aluno": 10,
    "SITUAÇÃO DO ALUNO": 10,
    "ATIVO": 4.5,
    "TRANSFERIDO": 4.5,
    "DESISTENTE": 4.5
}

LARGURAS_COLUNAS_ABAS_DISC = {
    "Nº": 1,
    "Nome do Aluno": 15,
    "1º BIM": 3,
    "2º BIM": 3,
    "3º BIM": 3,
    "4º BIM": 3,
    "NF": 3,
    "MG": 3,
    "MF": 3,
    "SITUAÇÃO DO ALUNO": 10,
    "PF": 3,
    "SF": 3
}

# Configurações de largura das colunas da aba SEC (em número de caracteres)
LARGURAS_COLUNAS_SEC_LETRAS = {
    "Nº": 5,
    "Nome do Aluno": 40,
    "ATIVO": 15,
    "TRANSFERIDO": 15,
    "DESISTENTE": 15,
    "SITUAÇÃO DO ALUNO": 40
}

# Configurações de largura das colunas das abas de disciplinas (em número de caracteres)
LARGURAS_COLUNAS_ABAS_DISC_LETRAS = {
    "Nº": 5,
    "Nome do Aluno": 60,
    "1º BIM": 10,
    "2º BIM": 10,
    "3º BIM": 10,
    "4º BIM": 10,
    "NF": 10,
    "MG": 10,
    "MF": 10,
    "SITUAÇÃO DO ALUNO": 40,
    "PF": 10,
    "SF": 10
}

# Nomes das tabelas Power Pivot
TBL_TURMAS_NAME = "tblTurmas"
TBL_ALUNOS_NAME = "tblAlunos"
TBL_NOTAS_NAME = "tblNotas"
TBL_DISCIPLINAS_NAME = "tblDisciplinas"

# Limites e posições
MAX_ALUNOS_FORMATAR = 50  # Número máximo de alunos a formatar por aba
LINHA_INICIAL = 5  # Linha inicial para dados na planilha
COLUNA_INICIAL = "A"  # Coluna inicial para dados

# Estilos e formatações
COR_ABA = "FFDAB9"  # Cor das abas (PeachPuff)

# Estilo de borda fina
BORDER_THIN = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# Preenchimentos (fills) para células
FILL_NOME_ALUNO = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Tomato
FILL_BIMESTRES = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # Orange
FILL_NOTA_FINAL = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")  # OrangeRed
FILL_SITUACAO = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")    # LemonChiffon

# Fontes e alinhamentos
FONTE_TITULO_TURMA = Font(name='Arial', size=14, bold=True, color="8B4513")  # SaddleBrown
ALINHAMENTO_CENTRALIZADO = Alignment(horizontal='center', vertical='center')

# Configurações dos dashboards
DASHBOARD_INDICADORES = [
    {
        "nome": "ALUNOS APROVADOS",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, ">=7")',
        "formato": None
    },
    {
        "nome": "ALUNOS REPROVADOS",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, "<7")',
        "formato": None
    },
    {
        "nome": "Nº ALUNOS COM MÉDIA > 8,0",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, ">=8")',
        "formato": None
    },
    {
        "nome": "Nº ALUNOS QUE NÃO ATINGIRAM MÉDIA > 8,0",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, "<8")',
        "formato": None
    },
    {
        "nome": "PERCENTUAL DE MÉDIAS > 5,0",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, ">=5")/COUNTA({col}{inicio}:{col}{fim})',
        "formato": '0.00%'
    },
    {
        "nome": "PERCENTUAL DE MÉDIAS < 5,0",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, "<5")/COUNTA({col}{inicio}:{col}{fim})',
        "formato": '0.00%'
    },
    {
        "nome": "MATRÍCULAS",
        "formula": lambda col, inicio, fim: f'=COUNTA({col}{inicio}:{col}{fim})',
        "formato": None
    },
    {
        "nome": "TAXA DE APROVAÇÃO (%)",
        "formula": lambda col, inicio, fim: f'=IF(COUNTA({col}{inicio}:{col}{fim})=0, 0, COUNTIF({col}{inicio}:{col}{fim}, ">=7")/COUNTA({col}{inicio}:{col}{fim}))',
        "formato": '0.00%'
    }
]

DASHBOARD_SEC_TURMA = [
    {
        "nome": "MATRÍCULAS",
        "formula": lambda col, inicio, fim: f'=COUNTA({col}{inicio}:{col}{fim})',
        "formato": None
    },
    {
        "nome": "ATIVOS",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, TRUE)',
        "formato": None
    },
    {
        "nome": "TRANSFERIDOS",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, TRUE)',
        "formato": None
    },
    {
        "nome": "DESISTENTES",
        "formula": lambda col, inicio, fim: f'=COUNTIF({col}{inicio}:{col}{fim}, TRUE)',
        "formato": None
    }
]

DASHBOARD_SEC_GERAL = [
    {
        "nome": "MATRÍCULAS",
        "formula": lambda refs: f'=SUM({",".join(refs)})',
        "formato": None
    },
    {
        "nome": "ATIVOS",
        "formula": lambda refs: f'=SUM({",".join(refs)})',
        "formato": None
    },
    {
        "nome": "TRANSFERIDOS",
        "formula": lambda refs: f'=SUM({",".join(refs)})',
        "formato": None
    },
    {
        "nome": "DESISTENTES",
        "formula": lambda refs: f'=SUM({",".join(refs)})',
        "formato": None
    },
    {
        "nome": "Nº ABANDONO(S)",
        "formula": lambda linha_atual: f'=K{linha_atual-1}',
        "formato": None
    },
    {
        "nome": "ABANDONO(S) (%)",
        "formula": lambda linha_atual: f'=K{linha_atual-1}/K{linha_atual-4}',
        "formato": '0.00%'
    }
]

DASHBOARD_SEC_APROVACAO = [
    {
        "nome": "TX APROVAÇÃO %",
        "formula": lambda col, inicio, fim: f'=AVERAGE({col}{inicio}:{col}{fim})',
        "formato": '0.00%'
    },
    {
        "nome": "TX REPROVAÇÃO %",
        "formula": lambda col, inicio, fim: f'=1-{col}{inicio-1}',
        "formato": '0.00%'
    }
]