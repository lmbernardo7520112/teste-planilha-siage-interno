import os
import json
import logging
import random
import re
from pathlib import Path
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.exceptions import IllegalCharacterError

try:
    from app.utils.excel_utils import (
        configurar_largura_colunas, criar_dashboard_turma, criar_dashboard_sec_turma,
        criar_dashboard_sec_geral, criar_dashboard_sec_aprovacao
    )
    from app.core.config import (
        COLUNAS, COLUNAS_SEC, DISCIPLINAS, DISCIPLINAS_NOMES, CAMINHO_IMAGEM, CAMINHO_PADRAO,
        NOME_ARQUIVO_PADRAO, LARGURAS_COLUNAS_SEC_LETRAS, LARGURAS_COLUNAS_ABAS_DISC_LETRAS,
        COR_ABA, FILL_NOME_ALUNO, FILL_BIMESTRES, FILL_NOTA_FINAL, FILL_SITUACAO,
        FONTE_TITULO_TURMA, ALINHAMENTO_CENTRALIZADO, CAMINHO_JSON,
        TBL_TURMAS_NAME, TBL_ALUNOS_NAME, TBL_DISCIPLINAS_NAME, TBL_NOTAS_NAME,
        BORDER_THIN, MAX_ALUNOS_FORMATAR
    )
except ImportError as e:
    print(f"Erro import: {e}. Tentando fallbacks...")
    try:
        from ..utils.excel_utils import *
        from ..core.config import *
    except ImportError:
        from utils.excel_utils import *
        from core.config import *

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_clean_student_name(name_string):
    if not isinstance(name_string, str):
        return str(name_string)
    return name_string.split('/')[0].strip()

def add_image_safely(ws, img_path, anchor):
    img_path_obj = Path(img_path)
    if not img_path_obj.exists():
        logger.warning(f"Imagem não encontrada: {img_path}.")
        return False
    try:
        img = Image(str(img_path_obj))
        ws.add_image(img, anchor)
        return True
    except Exception as e:
        logger.error(f"Erro add imagem {img_path}: {e}")
        return False

def criar_aba_em_branco(wb, titulo):
    ws = wb.create_sheet(title=titulo)
    ws.sheet_properties.tabColor = COR_ABA
    linha_atual = 1
    if add_image_safely(ws, CAMINHO_IMAGEM, 'A1'):
        try:
            img = Image(str(CAMINHO_IMAGEM))
            ws.row_dimensions[linha_atual].height = img.height * 0.75
        except:
            ws.row_dimensions[linha_atual].height = 50
        ws.merge_cells('A1:J1')
        cell_titulo_principal = ws['A1']
        cell_titulo_principal.alignment = ALINHAMENTO_CENTRALIZADO
    else:
        ws.merge_cells('A1:J1')
        cell_titulo_principal = ws['A1']
        cell_titulo_principal.alignment = ALINHAMENTO_CENTRALIZADO
        linha_atual += 1
    cell_titulo_principal.value = "ESCOLA COMPOSITOR LUIS RAMALHO"
    cell_titulo_principal.font = Font(name='Arial', size=22, bold=True, color="000080")
    linha_atual += 1
    ws.merge_cells(f'A{linha_atual}:J{linha_atual}')
    cell_subtitulo = ws[f'A{linha_atual}']
    cell_subtitulo.value = f"Relatório - {titulo}"
    cell_subtitulo.font = Font(name='Arial', size=12, italic=True)
    cell_subtitulo.alignment = ALINHAMENTO_CENTRALIZADO
    ws.row_dimensions[linha_atual].height = 20
    return ws

def criar_aba_sec(wb, turmas):
    """Cria a aba SEC com formatação completa das tabelas."""
    ws = wb.create_sheet(title="SEC")
    ws.sheet_properties.tabColor = COR_ABA
    linha_atual = 1
    max_linhas_tabela = MAX_ALUNOS_FORMATAR

    last_col_sec_letter = get_column_letter(len(COLUNAS_SEC))
    if add_image_safely(ws, CAMINHO_IMAGEM, 'A1'):
        try:
            img = Image(str(CAMINHO_IMAGEM))
            ws.row_dimensions[linha_atual].height = img.height * 0.75
        except:
            ws.row_dimensions[linha_atual].height = 50
        ws.merge_cells(f'A{linha_atual}:{last_col_sec_letter}{linha_atual}')
    else:
        ws.merge_cells(f'A{linha_atual}:{last_col_sec_letter}{linha_atual}')
    cell_titulo_principal = ws['A1']
    cell_titulo_principal.alignment = ALINHAMENTO_CENTRALIZADO
    cell_titulo_principal.value = "ESCOLA COMPOSITOR LUIS RAMALHO - Secretaria"
    cell_titulo_principal.font = Font(name='Arial', size=18, bold=True, color="000080")
    linha_atual += 2

    linhas_inicio_tabelas_headers = []
    num_alunos_por_turma = []
    colunas_largura_sec_completa = {**LARGURAS_COLUNAS_SEC_LETRAS, 'G': 25, 'H': 10, 'J': 25, 'K': 12, 'M': 20, 'N': 10, 'O': 10, 'P': 10, 'Q': 10, 'R': 15, 'S': 15, 'T': 15}
    configurar_largura_colunas(ws, colunas_largura_sec_completa)
    linhas_espaco_entre_turmas = 8

    for idx_turma, turma in enumerate(turmas):
        nome_turma = turma["nome_turma"]
        alunos = turma.get("alunos", [])
        num_alunos_turma = len(alunos)
        num_alunos_por_turma.append(num_alunos_turma)
        linha_inicio_bloco = 3 + (idx_turma * (1 + 1 + max_linhas_tabela + linhas_espaco_entre_turmas))
        linha_atual = linha_inicio_bloco

        cell_titulo_turma = ws.cell(row=linha_atual, column=1, value=nome_turma)
        cell_titulo_turma.font = FONTE_TITULO_TURMA
        cell_titulo_turma.alignment = ALINHAMENTO_CENTRALIZADO
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(COLUNAS_SEC))
        ws.row_dimensions[linha_atual].height = 30
        linha_atual += 1
        linha_header_turma_atual = linha_atual
        linhas_inicio_tabelas_headers.append(linha_header_turma_atual)

        # Cabeçalho da Tabela de Alunos
        for col_idx, col_nome in enumerate(COLUNAS_SEC, 1):
            cell = ws.cell(row=linha_atual, column=col_idx, value=col_nome)
            cell.border = BORDER_THIN
            cell.font = Font(bold=True)
            cell.alignment = ALINHAMENTO_CENTRALIZADO
            if col_nome == "Nome do Aluno":
                cell.fill = FILL_NOME_ALUNO
            elif col_nome == "SITUAÇÃO DO ALUNO":
                cell.fill = FILL_SITUACAO
            elif col_nome in ["ATIVO", "TRANSFERIDO", "DESISTENTE"]:
                cell.fill = FILL_BIMESTRES

        linha_inicio_dados = linha_atual + 1

        # Dados Alunos Reais
        alunos_processados = set()
        if alunos:
            for aluno in alunos:
                try:
                    num_aluno = int(aluno["numero"])
                    if not (0 < num_aluno <= max_linhas_tabela) or num_aluno in alunos_processados:
                        continue
                    alunos_processados.add(num_aluno)
                    linha_dados_atual = linha_inicio_dados + num_aluno - 1
                    ws.cell(row=linha_dados_atual, column=1, value=num_aluno).alignment = ALINHAMENTO_CENTRALIZADO
                    ws.cell(row=linha_dados_atual, column=2, value=aluno["nome"]).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws.cell(row=linha_dados_atual, column=3, value=True).alignment = ALINHAMENTO_CENTRALIZADO
                    ws.cell(row=linha_dados_atual, column=4, value=False).alignment = ALINHAMENTO_CENTRALIZADO
                    ws.cell(row=linha_dados_atual, column=5, value=False).alignment = ALINHAMENTO_CENTRALIZADO
                    ws.cell(row=linha_dados_atual, column=6, value=f'=IF(E{linha_dados_atual},"DESISTENTE",IF(D{linha_dados_atual},"TRANSFERIDO",IF(C{linha_dados_atual},"ATIVO","INDEFINIDO")))').alignment = ALINHAMENTO_CENTRALIZADO
                except Exception as e:
                    logger.error(f"Erro aluno {aluno.get('nome')} (SEC): {e}")
                    continue

        # Formatação Linhas Restantes
        for i in range(max_linhas_tabela):
            linha_formatar = linha_inicio_dados + i
            num_linha_atual = i + 1
            if num_linha_atual not in alunos_processados:
                ws.cell(row=linha_formatar, column=1, value=num_linha_atual).alignment = ALINHAMENTO_CENTRALIZADO
                ws.cell(row=linha_formatar, column=2).value = ""
                ws.cell(row=linha_formatar, column=3, value=False)
                ws.cell(row=linha_formatar, column=4, value=False)
                ws.cell(row=linha_formatar, column=5, value=False)
                ws.cell(row=linha_formatar, column=6, value=f'=IF(E{linha_formatar},"DESISTENTE",IF(D{linha_formatar},"TRANSFERIDO",IF(C{linha_formatar},"ATIVO","INDEFINIDO")))').alignment = ALINHAMENTO_CENTRALIZADO
            for col_idx in range(1, len(COLUNAS_SEC) + 1):
                ws.cell(row=linha_formatar, column=col_idx).border = BORDER_THIN

        # Dash Turma
        try:
            criar_dashboard_sec_turma(ws, linha_header_turma_atual, linha_inicio_dados, num_alunos_turma)
        except Exception as e:
            logger.error(f"Erro dash SEC turma {nome_turma}: {e}")

    # Dashboards Gerais
    if linhas_inicio_tabelas_headers:
        try:
            logger.info("Criando dashboard SEC Geral...")
            criar_dashboard_sec_geral(ws, linhas_inicio_tabelas_headers, num_alunos_por_turma)
        except Exception as e:
            logger.error(f"Erro dash SEC Geral: {e}")
        try:
            logger.info("Criando dashboard SEC Aprovação...")
            criar_dashboard_sec_aprovacao(ws, turmas, linhas_inicio_tabelas_headers)
        except Exception as e:
            logger.error(f"Erro dash SEC Aprovação: {e}")
    return ws

def criar_aba_disciplina(wb, titulo_disciplina, turmas):
    ws = wb.create_sheet(title=titulo_disciplina)
    ws.sheet_properties.tabColor = COR_ABA
    linha_atual = 1
    max_linhas_tabela = MAX_ALUNOS_FORMATAR
    last_col_disc_letter = get_column_letter(len(COLUNAS))
    if add_image_safely(ws, CAMINHO_IMAGEM, 'A1'):
        try:
            img = Image(str(CAMINHO_IMAGEM))
            ws.row_dimensions[linha_atual].height = img.height * 0.75
        except:
            ws.row_dimensions[linha_atual].height = 50
        ws.merge_cells(f'A{linha_atual}:{last_col_disc_letter}{linha_atual}')
    else:
        ws.merge_cells(f'A{linha_atual}:{last_col_disc_letter}{linha_atual}')
    cell_titulo_principal = ws['A1']
    cell_titulo_principal.alignment = ALINHAMENTO_CENTRALIZADO
    cell_titulo_principal.value = f"ESCOLA COMPOSITOR LUIS RAMALHO - {titulo_disciplina}"
    cell_titulo_principal.font = Font(name='Arial', size=18, bold=True, color="000080")
    linha_atual += 2
    colunas_largura_disc_completa = {**LARGURAS_COLUNAS_ABAS_DISC_LETRAS, 'N': 30, 'O': 12, 'P': 12, 'Q': 12, 'R': 12}
    configurar_largura_colunas(ws, colunas_largura_disc_completa)
    linhas_espaco_entre_turmas_disc = 15
    linhas_por_bloco_turma_disc = 1 + 1 + max_linhas_tabela + linhas_espaco_entre_turmas_disc
    for idx_turma, turma in enumerate(turmas):
        nome_turma = turma["nome_turma"]
        alunos = turma.get("alunos", [])
        num_alunos_turma = len(alunos)
        linha_inicio_bloco = 3 + (idx_turma * linhas_por_bloco_turma_disc)
        linha_atual = linha_inicio_bloco
        cell_titulo_turma = ws.cell(row=linha_atual, column=1, value=f"{nome_turma}")
        cell_titulo_turma.font = FONTE_TITULO_TURMA
        cell_titulo_turma.alignment = ALINHAMENTO_CENTRALIZADO
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(COLUNAS))
        ws.row_dimensions[linha_atual].height = 30
        linha_atual += 1
        linha_header_tabela = linha_atual
        for col_idx, col_nome in enumerate(COLUNAS, 1):
            cell = ws.cell(row=linha_atual, column=col_idx, value=col_nome)
            cell.border = BORDER_THIN
            cell.font = Font(bold=True)
            cell.alignment = ALINHAMENTO_CENTRALIZADO
            if col_nome == "Nome do Aluno":
                cell.fill = FILL_NOME_ALUNO
            elif "BIM" in col_nome:
                cell.fill = FILL_BIMESTRES
            elif col_nome in ["NF", "MG", "MF", "PF", "SF"]:
                cell.fill = FILL_NOTA_FINAL
            elif col_nome == "SITUAÇÃO DO ALUNO":
                cell.fill = FILL_SITUACAO
        linha_inicio_dados = linha_atual + 1
        alunos_processados_disc = set()
        if alunos:
            for aluno in alunos:
                try:
                    num_aluno = int(aluno["numero"])
                    if not (0 < num_aluno <= max_linhas_tabela) or num_aluno in alunos_processados_disc:
                        continue
                    alunos_processados_disc.add(num_aluno)
                    linha_dados_atual = linha_inicio_dados + num_aluno - 1
                    ws.cell(row=linha_dados_atual, column=1, value=num_aluno).alignment = ALINHAMENTO_CENTRALIZADO
                    ws.cell(row=linha_dados_atual, column=2, value=aluno["nome"]).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    for col_idx in range(3, 7):
                        ws.cell(row=linha_dados_atual, column=col_idx, value=round(random.uniform(3.0, 10.0), 2)).number_format = '0.00'
                        ws.cell(row=linha_dados_atual, column=col_idx).alignment = ALINHAMENTO_CENTRALIZADO
                except Exception as e:
                    logger.error(f"Erro aluno {aluno.get('nome')} ({titulo_disciplina}): {e}")
        for i in range(max_linhas_tabela):
            row_formula = linha_inicio_dados + i
            num_linha_atual = i + 1
            aluno_presente = num_linha_atual in alunos_processados_disc
            if not aluno_presente:
                ws.cell(row=row_formula, column=1, value=num_linha_atual).alignment = ALINHAMENTO_CENTRALIZADO
                ws.cell(row=row_formula, column=2).value = ""
            for col_idx in range(3, 7):
                ws.cell(row=row_formula, column=col_idx).value = ""
            ws.cell(row=row_formula, column=7).value = f'=IF(COUNT(C{row_formula}:F{row_formula})>0, IFERROR(AVERAGE(C{row_formula}:F{row_formula}), ""), "")'
            ws.cell(row=row_formula, column=8).value = f'=IF(COUNT(C{row_formula}:F{row_formula})>0, IFERROR(SUM(C{row_formula}:F{row_formula})/4, ""), "")'
            ws.cell(row=row_formula, column=9).value = f'=IF(ISNUMBER(H{row_formula}),IF(H{row_formula}<7, IFERROR((0.6*H{row_formula})+(0.4*G{row_formula}), "-"), "-"),"")'
            ws.cell(row=row_formula, column=10).value = f'=IF(ISNUMBER(H{row_formula}),IF(H{row_formula}<2.5, "REPROVADO", IF(H{row_formula}<7, "FINAL", "APROVADO")),"")'
            ws.cell(row=row_formula, column=11).value = f'=IF(ISNUMBER(H{row_formula}),IF(H{row_formula}<7, IFERROR(MAX(0, 12.5-(1.5*H{row_formula})), "-"), "-"),"")'
            ws.cell(row=row_formula, column=12).value = f'=IF(K{row_formula}="-", IF(J{row_formula}="APROVADO","APROVADO",""), IF(AND(ISNUMBER(K{row_formula}), ISNUMBER(G{row_formula})), IF(G{row_formula}>=K{row_formula}, "APROVADO FINAL", "REPROVADO FINAL"),""))'
            for col_idx in range(1, len(COLUNAS) + 1):
                cell = ws.cell(row=row_formula, column=col_idx)
                cell.border = BORDER_THIN
                if col_idx >= 7 and col_idx != 10 and col_idx != 12:
                    cell.number_format = '0.00'
                if col_idx >= 3:
                    cell.alignment = ALINHAMENTO_CENTRALIZADO
        try:
            criar_dashboard_turma(ws, linha_header_tabela, linha_inicio_dados, num_alunos_turma)
        except Exception as e:
            logger.error(f"Erro dash turma {nome_turma} ({titulo_disciplina}): {e}")
    return ws

def criar_aba_boletim(wb, turmas):
    ws = wb.create_sheet(title="BOLETIM")
    ws.sheet_properties.tabColor = COR_ABA
    linha_atual = 1
    max_linhas_tabela = MAX_ALUNOS_FORMATAR
    headers = ["Nº", "ALUNO"]
    col_widths = {'A': 5, 'B': 35}
    disciplina_start_cols = {}
    current_col_idx = 3
    for disciplina in DISCIPLINAS:
        disciplina_start_cols[disciplina] = current_col_idx
        headers.extend([f"B1", f"B2", f"B3", f"B4", f"NF", f"MG"])
        for i in range(6):
            col_widths[get_column_letter(current_col_idx + i)] = 6
        current_col_idx += 6
    last_col_idx = current_col_idx - 1
    last_col_letter = get_column_letter(last_col_idx)
    if add_image_safely(ws, CAMINHO_IMAGEM, 'A1'):
        try:
            img = Image(str(CAMINHO_IMAGEM))
            ws.row_dimensions[linha_atual].height = img.height * 0.75
        except:
            ws.row_dimensions[linha_atual].height = 50
        ws.merge_cells(f'A{linha_atual}:{last_col_letter}{linha_atual}')
    else:
        ws.merge_cells(f'A{linha_atual}:{last_col_letter}{linha_atual}')
    cell_titulo_principal = ws['A1']
    cell_titulo_principal.alignment = ALINHAMENTO_CENTRALIZADO
    cell_titulo_principal.value = "ESCOLA COMPOSITOR LUIS RAMALHO - BOLETIM GERAL"
    cell_titulo_principal.font = Font(name='Arial', size=18, bold=True, color="000080")
    linha_atual += 2
    configurar_largura_colunas(ws, col_widths)
    linhas_espaco_entre_turmas_bol = 5
    for idx_turma, turma in enumerate(turmas):
        nome_turma = turma["nome_turma"]
        alunos = turma.get("alunos", [])
        num_alunos_turma = len(alunos)
        linha_inicio_bloco = 3 + (idx_turma * (1 + 1 + 1 + max_linhas_tabela + linhas_espaco_entre_turmas_bol))
        linha_atual = linha_inicio_bloco
        cell_titulo_turma = ws.cell(row=linha_atual, column=1, value=f"{nome_turma} - BOLETIM")
        cell_titulo_turma.font = FONTE_TITULO_TURMA
        cell_titulo_turma.alignment = ALINHAMENTO_CENTRALIZADO
        ws.merge_cells(f'A{linha_atual}:{last_col_letter}{linha_atual}')
        ws.row_dimensions[linha_atual].height = 30
        linha_atual += 1
        ws.cell(row=linha_atual, column=1).border = BORDER_THIN
        ws.cell(row=linha_atual, column=2).border = BORDER_THIN
        for disciplina, start_col in disciplina_start_cols.items():
            end_col = start_col + 5
            ws.merge_cells(start_row=linha_atual, start_column=start_col, end_row=linha_atual, end_column=end_col)
            cell_disc_header = ws.cell(row=linha_atual, column=start_col, value=disciplina)
            cell_disc_header.font = Font(bold=True, size=8)
            cell_disc_header.alignment = ALINHAMENTO_CENTRALIZADO
            cell_disc_header.fill = FILL_BIMESTRES
            for c_idx in range(start_col, end_col + 1):
                ws.cell(row=linha_atual, column=c_idx).border = BORDER_THIN
        linha_atual += 1
        linha_header_boletim = linha_atual
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=linha_atual, column=col_idx, value=header_text)
            cell.border = BORDER_THIN
            cell.font = Font(bold=True, size=7)
            cell.alignment = ALINHAMENTO_CENTRALIZADO
            if header_text == "ALUNO":
                cell.fill = FILL_NOME_ALUNO
            elif "B" in header_text:
                cell.fill = FILL_BIMESTRES
            elif header_text in ["NF", "MG"]:
                cell.fill = FILL_NOTA_FINAL
        linha_atual += 1
        linha_inicio_dados = linha_atual
        alunos_processados_bol = set()
        if alunos:
            for aluno in alunos:
                try:
                    num_aluno = int(aluno["numero"])
                    if not (0 < num_aluno <= max_linhas_tabela) or num_aluno in alunos_processados_bol:
                        continue
                    alunos_processados_bol.add(num_aluno)
                    linha_dados_atual = linha_inicio_dados + num_aluno - 1
                    nome_aluno_completo = aluno["nome"]
                    nome_aluno_limpo = get_clean_student_name(nome_aluno_completo)
                    ws.cell(row=linha_dados_atual, column=1, value=num_aluno).alignment = ALINHAMENTO_CENTRALIZADO
                    ws.cell(row=linha_dados_atual, column=2, value=nome_aluno_completo).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    for disciplina, start_col in disciplina_start_cols.items():
                        if disciplina not in wb.sheetnames:
                            logger.error(f"Aba '{disciplina}' não existe (Boletim)")
                            continue
                        ws_disciplina = wb[disciplina]
                        linha_ref = _get_student_row_in_discipline_sheet(ws_disciplina, nome_turma, nome_aluno_limpo, turmas)
                        col_letters_disc = ['C', 'D', 'E', 'F', 'G', 'H']
                        for i, col_let_disc in enumerate(col_letters_disc):
                            cell_boletim = ws.cell(row=linha_dados_atual, column=start_col + i)
                            if linha_ref:
                                cell_boletim.value = f"='{disciplina}'!{col_let_disc}{linha_ref}"
                                cell_boletim.number_format = '0.00'
                            else:
                                cell_boletim.value = 0.00
                                cell_boletim.number_format = '0.00'
                            cell_boletim.alignment = ALINHAMENTO_CENTRALIZADO
                except Exception as e:
                    logger.error(f"Erro aluno {aluno.get('nome')} (Boletim): {e}")
        for i in range(max_linhas_tabela):
            linha_formatar = linha_inicio_dados + i
            num_linha_atual = i + 1
            if num_linha_atual not in alunos_processados_bol:
                ws.cell(row=linha_formatar, column=1, value=num_linha_atual).alignment = ALINHAMENTO_CENTRALIZADO
                ws.cell(row=linha_formatar, column=2).value = ""
                for disciplina, start_col in disciplina_start_cols.items():
                    for col_offset in range(6):
                        ws.cell(row=linha_formatar, column=start_col + col_offset).value = ""
                for col_idx in range(1, last_col_idx + 1):
                    ws.cell(row=linha_formatar, column=col_idx).border = BORDER_THIN
    return ws

def _get_student_row_in_discipline_sheet(ws_disciplina, turma_nome, aluno_nome_limpo, turmas_list):
    try:
        turma_index = next(i for i, t in enumerate(turmas_list) if t['nome_turma'] == turma_nome)
        aluno_obj = next(a for a in turmas_list[turma_index]['alunos'] if get_clean_student_name(a['nome']) == aluno_nome_limpo)
        aluno_numero = int(aluno_obj['numero'])
        linhas_espaco_entre_turmas_disc = 15
        linhas_por_bloco_turma_disc = 1 + 1 + MAX_ALUNOS_FORMATAR + linhas_espaco_entre_turmas_disc
        linha_inicio_bloco = 3 + (turma_index * linhas_por_bloco_turma_disc)
        linha_header_tabela = linha_inicio_bloco + 1
        linha_inicio_dados = linha_header_tabela + 1
        linha_aluno_calculada = linha_inicio_dados + aluno_numero - 1
        if not (0 < aluno_numero <= MAX_ALUNOS_FORMATAR):
            return None
        nome_na_celula = ws_disciplina.cell(row=linha_aluno_calculada, column=2).value
        if nome_na_celula and get_clean_student_name(str(nome_na_celula)) == aluno_nome_limpo:
            return linha_aluno_calculada
        else:
            logger.warning(f"Lookup: Calc={linha_aluno_calculada} para '{aluno_nome_limpo}', Encont='{get_clean_student_name(str(nome_na_celula))}' em {ws_disciplina.title}/{turma_nome}. Buscando...")
        for offset in range(-3, 4):
            row_num_check = linha_aluno_calculada + offset
            if row_num_check < linha_inicio_dados or row_num_check >= linha_inicio_dados + MAX_ALUNOS_FORMATAR:
                continue
            cell_value = ws_disciplina.cell(row=row_num_check, column=2).value
            if cell_value and get_clean_student_name(str(cell_value)) == aluno_nome_limpo:
                logger.info(f"Lookup: Aluno {aluno_nome_limpo} linha {row_num_check} (busca).")
                return row_num_check
        logger.error(f"Lookup: Aluno {aluno_nome_limpo} NÃO encontrado em {ws_disciplina.title}/{turma_nome}.")
        return None
    except Exception as e:
        logger.error(f"Lookup: Erro {aluno_nome_limpo}/{turma_nome}: {e}")
        return None

def criar_aba_dashboard_powerpivot(wb, turmas_list):
    logger.info("Criando aba DASHBOARD para Power Pivot")
    ws = wb.create_sheet(title="DASHBOARD")
    ws.sheet_properties.tabColor = "ADD8E6"
    turmas_data, alunos_data, disciplinas_data, notas_data = [], [], [], []
    turma_id_map, aluno_id_map = {}, {}
    current_turma_id, current_aluno_id, current_nota_id = 1, 1, 1
    turmas_headers = ["TurmaID", "NomeTurma", "GradeLevel", "ClassLetter"]
    for turma in turmas_list:
        nome_turma = turma["nome_turma"]
        turma_id = f"T{current_turma_id:02d}"
        turma_id_map[nome_turma] = turma_id
        current_turma_id += 1
        grade_level, class_letter = "N/D", "N/D"
        match = re.match(r"(\d+º)\s*ANO\s*([A-Z])", nome_turma, re.IGNORECASE)
        if match:
            grade_level, class_letter = match.group(1), match.group(2).upper()
        turmas_data.append([turma_id, nome_turma, grade_level, class_letter])
    alunos_headers = ["StudentID", "Name", "Status", "TurmaID"]
    for turma in turmas_list:
        nome_turma = turma["nome_turma"]
        turma_id = turma_id_map.get(nome_turma)
        if not turma_id:
            continue
        for aluno in turma.get("alunos", []):
            try:
                student_id = f"A{current_aluno_id:04d}"
                nome_aluno_limpo = get_clean_student_name(aluno["nome"])
                aluno_id_map[(nome_turma, nome_aluno_limpo)] = student_id
                current_aluno_id += 1
                status = "ATIVO"
                alunos_data.append([student_id, nome_aluno_limpo, status, turma_id])
            except Exception as e:
                logger.error(f"Erro Tabela Alunos (Dash) {aluno.get('nome')}: {e}")
    disciplinas_headers = ["DisciplineCode", "DisciplineName"]
    for code in DISCIPLINAS:
        disciplinas_data.append([code, DISCIPLINAS_NOMES.get(code, code)])
    notas_headers = ["NotaID", "StudentID", "DisciplineCode", "BIM1Grade", "BIM2Grade", "BIM3Grade", "BIM4Grade", "NF", "MG", "MF", "Situation"]
    col_map_disc_to_notas = {'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8, 'I': 9, 'J': 10}
    for turma in turmas_list:
        nome_turma = turma["nome_turma"]
        for aluno in turma.get("alunos", []):
            nome_aluno_limpo = get_clean_student_name(aluno["nome"])
            student_id = aluno_id_map.get((nome_turma, nome_aluno_limpo))
            if not student_id:
                continue
            for discipline_code in DISCIPLINAS:
                ws_disc = wb.get_sheet_by_name(discipline_code) if discipline_code in wb.sheetnames else None
                if not ws_disc:
                    continue
                linha_aluno = _get_student_row_in_discipline_sheet(ws_disc, nome_turma, nome_aluno_limpo, turmas_list)
                if linha_aluno:
                    try:
                        nota_row = [f"N{current_nota_id:05d}", student_id, discipline_code] + [None] * (len(notas_headers) - 3)
                        valid_data = False
                        for col_letter, header_idx in col_map_disc_to_notas.items():
                            col_idx_num = column_index_from_string(col_letter)
                            cell_value = ws_disc.cell(row=linha_aluno, column=col_idx_num).value
                            processed = None
                            if header_idx <= 9:
                                try:
                                    if isinstance(cell_value, (int, float)):
                                        processed = float(cell_value)
                                        valid_data = True
                                    elif isinstance(cell_value, str) and cell_value == '-':
                                        processed = None
                                    else:
                                        processed = float(cell_value)
                                        valid_data = True
                                except:
                                    processed = None
                            elif header_idx == 10:
                                processed = str(cell_value).strip() if cell_value else "N/D"
                                valid_data = True
                            nota_row[header_idx] = processed
                        if valid_data:
                            notas_data.append(nota_row)
                            current_nota_id += 1
                    except Exception as read_err:
                        logger.error(f"Erro leitura notas (Dash) {nome_aluno_limpo}/{discipline_code}: {read_err}")
    current_row = 1
    table_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    def write_table(ws_target, start_row, headers, data, table_name, style_info):
        ws_target.cell(row=start_row, column=1, value=table_name.replace("tbl", "Tabela ")).font = Font(bold=True, size=12)
        current_r = start_row + 1
        ws_target.append(headers)
        header_row = current_r
        current_r += 1
        if not data:
            ws_target.cell(row=current_r, column=1, value="Nenhum dado disponível.")
            current_r += 1
        else:
            for row_data in data:
                try:
                    safe_row = [re.sub(r'[\x00-\x1F\x7F]', '', str(item)) if isinstance(item, str) else (item if item is not None else "") for item in row_data]
                    ws_target.append(safe_row)
                    current_r += 1
                except Exception as write_err:
                    logger.error(f"Erro escrita {table_name}: {row_data} -> {write_err}.")
        data_end_row = current_r - 1
        if data_end_row >= header_row:
            table_ref = f"A{header_row}:{get_column_letter(len(headers))}{data_end_row}"
            try:
                excel_table = Table(displayName=table_name, ref=table_ref)
                excel_table.tableStyleInfo = style_info
                ws_target.add_table(excel_table)
                logger.info(f"Tabela '{table_name}' criada: {table_ref}")
            except Exception as table_err:
                logger.error(f"Erro add tabela '{table_name}' ref '{table_ref}': {table_err}")
        else:
            logger.warning(f"Tabela '{table_name}' não criada (sem dados).")
        ws_target.column_dimensions['A'].width = 15
        ws_target.column_dimensions['B'].width = 35
        return current_r + 1
    logger.info("Escrevendo tabelas na aba DASHBOARD...")
    current_row = write_table(ws, current_row, turmas_headers, turmas_data, TBL_TURMAS_NAME, table_style)
    current_row = write_table(ws, current_row, alunos_headers, alunos_data, TBL_ALUNOS_NAME, table_style)
    current_row = write_table(ws, current_row, disciplinas_headers, disciplinas_data, TBL_DISCIPLINAS_NAME, table_style)
    current_row = write_table(ws, current_row, notas_headers, notas_data, TBL_NOTAS_NAME, table_style)
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 10
    ws.column_dimensions['K'].width = 18
    logger.info("Aba DASHBOARD criada.")
    return ws

def criar_planilha():
    logger.info("="*45)
    logger.info(" Iniciando criação da planilha ")
    logger.info("="*45)
    if not CAMINHO_JSON.exists():
        logger.error(f"JSON não encontrado: {CAMINHO_JSON}")
        raise FileNotFoundError(f"Arquivo JSON '{CAMINHO_JSON}' não encontrado.")
    wb = Workbook()
    wb.remove(wb.active)
    try:
        with open(CAMINHO_JSON, 'r', encoding='utf-8') as f:
            dados = json.load(f)
            turmas = dados.get("turmas")
        if not turmas:
            raise ValueError("JSON inválido ou vazio.")
        logger.info(f"JSON carregado. {len(turmas)} turmas.")
    except Exception as load_err:
        logger.error(f"Erro carregando JSON: {load_err}")
        raise

    logger.info("-> Criando aba SEC...")
    criar_aba_sec(wb, turmas)
    logger.info("-> Criando abas de Disciplinas...")
    for disciplina in DISCIPLINAS:
        logger.info(f"   - Criando aba {disciplina}...")
        criar_aba_disciplina(wb, disciplina, turmas)
    logger.info("-> Criando aba BOLETIM...")
    criar_aba_boletim(wb, turmas)
    logger.info("-> Criando aba DASHBOARD...")
    criar_aba_dashboard_powerpivot(wb, turmas)
    abas_adicionais = ["INDIVIDUAL", "BOL", "RESULTADO", "FREQUÊNCIA"]
    logger.info(f"-> Criando abas em branco: {', '.join(abas_adicionais)}")
    for aba in abas_adicionais:
        criar_aba_em_branco(wb, aba)

    diretorio_saida = Path(CAMINHO_PADRAO)
    try:
        diretorio_saida.mkdir(parents=True, exist_ok=True)
    except OSError as dir_err:
        logger.error(f"Erro criando diretório '{diretorio_saida}': {dir_err}")
        raise
    caminho_completo = diretorio_saida / NOME_ARQUIVO_PADRAO
    logger.info(f"-> Tentando salvar em: {caminho_completo}")
    try:
        wb.save(caminho_completo)
        logger.info("="*45)
        logger.info(f" Planilha salva: {caminho_completo}")
        logger.info("="*45)
        return str(caminho_completo)
    except Exception as save_err:
        logger.error(f"Falha ao salvar: {save_err}")
        logger.exception("Detalhes:")
        raise

if __name__ == '__main__':
    print(f"Executando {__file__}...")
    print(f"JSON: {CAMINHO_JSON}")
    print(f"Imagem: {CAMINHO_IMAGEM}")
    print(f"Saída: {Path(CAMINHO_PADRAO) / NOME_ARQUIVO_PADRAO}")
    try:
        criar_planilha()
        print("\nExecução concluída!")
    except Exception as main_err:
        print(f"\nERRO: {main_err}")
        logger.exception("Detalhes:")