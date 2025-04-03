import logging
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties

try:
    from app.core.config import (
        DASHBOARD_INDICADORES, FILL_BIMESTRES, DASHBOARD_SEC_TURMA,
        DASHBOARD_SEC_GERAL, ALINHAMENTO_CENTRALIZADO, DASHBOARD_SEC_APROVACAO,
        DISCIPLINAS, BORDER_THIN, MAX_ALUNOS_FORMATAR
    )
except ImportError:
    from core.config import (
        DASHBOARD_INDICADORES, FILL_BIMESTRES, DASHBOARD_SEC_TURMA,
        DASHBOARD_SEC_GERAL, ALINHAMENTO_CENTRALIZADO, DASHBOARD_SEC_APROVACAO,
        DISCIPLINAS, BORDER_THIN, MAX_ALUNOS_FORMATAR
    )

logger = logging.getLogger(__name__)

def configurar_largura_colunas(ws, colunas_largura_letras):
    """Configura a largura das colunas com base em {Letra: Largura}."""
    for col_letra, largura_unidades in colunas_largura_letras.items():
        try:
            ws.column_dimensions[col_letra.upper()].width = largura_unidades
        except Exception as e:
            logger.error(f"Erro largura coluna '{col_letra}' aba {ws.title}: {e}")

def criar_dashboard_turma(ws, linha_inicio_tabela_header, linha_inicio_dados, num_alunos):
    """Cria o dashboard de resumo por turma na aba de disciplina."""
    dashboard_col_start_idx = 14
    num_bimestres = 4
    ws_title = ws.title
    dashboard_linha = linha_inicio_tabela_header + 1
    cell_titulo = ws.cell(row=dashboard_linha, column=dashboard_col_start_idx, value="Resumo da Turma")
    cell_titulo.font = Font(bold=True, size=12)
    cell_titulo.alignment = ALINHAMENTO_CENTRALIZADO
    ws.merge_cells(start_row=dashboard_linha, start_column=dashboard_col_start_idx, end_row=dashboard_linha, end_column=dashboard_col_start_idx + num_bimestres)
    dashboard_linha += 1
    ws.cell(row=dashboard_linha, column=dashboard_col_start_idx, value="Indicador").font = Font(bold=True)
    ws.cell(row=dashboard_linha, column=dashboard_col_start_idx).alignment = ALINHAMENTO_CENTRALIZADO
    for i in range(num_bimestres):
        cell = ws.cell(row=dashboard_linha, column=dashboard_col_start_idx + 1 + i, value=f"{i+1}º Bimestre")
        cell.font = Font(bold=True)
        cell.alignment = ALINHAMENTO_CENTRALIZADO
        cell.fill = FILL_BIMESTRES
    inicio = linha_inicio_dados
    fim = linha_inicio_dados + num_alunos - 1 if num_alunos > 0 else inicio
    bimestre_cols_letters = ['C', 'D', 'E', 'F']
    indicadores_inteiros = ["ALUNOS APROVADOS", "ALUNOS REPROVADOS", "Nº ALUNOS COM MÉDIA > 8,0", "Nº ALUNOS QUE NÃO ATINGIRAM MÉDIA > 8,0", "MATRÍCULAS"]
    linha_final_indicadores = dashboard_linha
    for indicador in DASHBOARD_INDICADORES:
        dashboard_linha += 1
        linha_final_indicadores = dashboard_linha
        ws.cell(row=dashboard_linha, column=dashboard_col_start_idx, value=indicador["nome"]).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if callable(indicador.get("formula")):
            try:
                for i, bim_col_letter in enumerate(bimestre_cols_letters):
                    cell_valor = ws.cell(row=dashboard_linha, column=dashboard_col_start_idx + 1 + i)
                    formula_str = indicador["formula"](bim_col_letter, inicio, fim, ws_title)
                    cell_valor.value = formula_str
                    cell_valor.alignment = ALINHAMENTO_CENTRALIZADO
                    fmt = indicador.get("formato")
                    is_int = indicador["nome"] in indicadores_inteiros
                    if fmt:
                        cell_valor.number_format = fmt
                    elif is_int:
                        cell_valor.number_format = '0'
                    else:
                        cell_valor.number_format = '0.00'
            except Exception as e:
                logger.error(f"Erro formula '{indicador['nome']}' ({ws.title}): {e}")
        else:
            logger.warning(f"Formula não chamável '{indicador['nome']}' ({ws.title})")
    max_col_dashboard = dashboard_col_start_idx + num_bimestres
    for row in range(linha_inicio_tabela_header + 1, linha_final_indicadores + 1):
        for col in range(dashboard_col_start_idx, max_col_dashboard + 1):
            ws.cell(row=row, column=col).border = BORDER_THIN

def criar_dashboard_sec_turma(ws, linha_inicio_tabela_header, linha_inicio_dados, num_alunos):
    """Cria o dashboard de resumo de status por turma na aba SEC."""
    dashboard_col_start_idx = 7
    dashboard_col_end_idx = 8
    dashboard_linha = linha_inicio_tabela_header + 1
    cell_titulo = ws.cell(row=dashboard_linha, column=dashboard_col_start_idx, value="Resumo Status Turma")
    cell_titulo.font = Font(bold=True, size=12)
    cell_titulo.alignment = ALINHAMENTO_CENTRALIZADO
    ws.merge_cells(start_row=dashboard_linha, start_column=dashboard_col_start_idx, end_row=dashboard_linha, end_column=dashboard_col_end_idx)
    dashboard_linha += 1
    ws.cell(row=dashboard_linha, column=dashboard_col_start_idx, value="Indicador").font = Font(bold=True)
    ws.cell(row=dashboard_linha, column=dashboard_col_start_idx).alignment = ALINHAMENTO_CENTRALIZADO
    ws.cell(row=dashboard_linha, column=dashboard_col_end_idx, value="Qtd").font = Font(bold=True)
    ws.cell(row=dashboard_linha, column=dashboard_col_end_idx).alignment = ALINHAMENTO_CENTRALIZADO
    inicio = linha_inicio_dados
    fim = linha_inicio_dados + num_alunos - 1 if num_alunos > 0 else inicio
    indicadores_inteiros = ["MATRÍCULAS", "ATIVOS", "TRANSFERIDOS", "DESISTENTES"]
    linha_final_indicadores_sec_turma = dashboard_linha
    for indicador in DASHBOARD_SEC_TURMA:
        dashboard_linha += 1
        linha_final_indicadores_sec_turma = dashboard_linha
        ws.cell(row=dashboard_linha, column=dashboard_col_start_idx, value=indicador["nome"]).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_valor = ws.cell(row=dashboard_linha, column=dashboard_col_end_idx)
        if callable(indicador.get("formula")):
            try:
                formula_str = indicador["formula"]('C', inicio, fim)
                cell_valor.value = formula_str
                cell_valor.alignment = ALINHAMENTO_CENTRALIZADO
                if indicador["nome"] in indicadores_inteiros:
                    cell_valor.number_format = '0'
                else:
                    cell_valor.number_format = '0.00'
            except Exception as e:
                logger.error(f"Erro formula SEC_TURMA '{indicador['nome']}': {e}")
                cell_valor.value = "#ERRO!"
        else:
            logger.warning(f"Formula não chamável SEC_TURMA '{indicador['nome']}'")
            cell_valor.value = "#N/A"
    for row in range(linha_inicio_tabela_header + 1, linha_final_indicadores_sec_turma + 1):
        for col in range(dashboard_col_start_idx, dashboard_col_end_idx + 1):
            ws.cell(row=row, column=col).border = BORDER_THIN

def criar_dashboard_sec_geral(ws, linhas_inicio_tabelas_headers, num_alunos_por_turma, start_col='L', start_row=4):
    """Cria o dashboard geral e o GRÁFICO DE STATUS na aba SEC."""
    if not linhas_inicio_tabelas_headers:
        logger.warning("Sem dados para dashboard SEC GERAL.")
        return
    dashboard_col_start_idx = column_index_from_string(start_col)
    dashboard_col_end_idx = dashboard_col_start_idx + 1
    dashboard_linha_ref = start_row
    cell_titulo = ws.cell(row=dashboard_linha_ref, column=dashboard_col_start_idx, value="Resumo Geral Escola")
    cell_titulo.font = Font(bold=True, size=12)
    cell_titulo.alignment = ALINHAMENTO_CENTRALIZADO
    ws.merge_cells(start_row=dashboard_linha_ref, start_column=dashboard_col_start_idx, end_row=dashboard_linha_ref, end_column=dashboard_col_end_idx)
    dashboard_linha_atual = dashboard_linha_ref + 1
    ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx, value="Indicador Geral").font = Font(bold=True)
    ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx).alignment = ALINHAMENTO_CENTRALIZADO
    ws.cell(row=dashboard_linha_atual, column=dashboard_col_end_idx, value="Total").font = Font(bold=True)
    ws.cell(row=dashboard_linha_atual, column=dashboard_col_end_idx).alignment = ALINHAMENTO_CENTRALIZADO
    indicador_row_offset = 2
    matriculas_refs = [f'H{lh + indicador_row_offset + 1}' for lh in linhas_inicio_tabelas_headers]
    ativos_refs = [f'H{lh + indicador_row_offset + 2}' for lh in linhas_inicio_tabelas_headers]
    transferidos_refs = [f'H{lh + indicador_row_offset + 3}' for lh in linhas_inicio_tabelas_headers]
    desistentes_refs = [f'H{lh + indicador_row_offset + 4}' for lh in linhas_inicio_tabelas_headers]
    refs_por_indicador = {"MATRÍCULAS": matriculas_refs, "ATIVOS": ativos_refs, "TRANSFERIDOS": transferidos_refs, "DESISTENTES": desistentes_refs}
    indicadores_inteiros = ["MATRÍCULAS", "ATIVOS", "TRANSFERIDOS", "DESISTENTES", "Nº ABANDONO(S)"]
    valor_col_letter = get_column_letter(dashboard_col_end_idx)
    linha_inicio_dados_geral = dashboard_linha_atual + 1
    for indicador in DASHBOARD_SEC_GERAL:
        dashboard_linha_atual += 1
        ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx, value=indicador["nome"]).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_valor = ws.cell(row=dashboard_linha_atual, column=dashboard_col_end_idx)
        formula_aplicada = False
        if callable(indicador.get("formula")):
            try:
                if indicador["nome"] in refs_por_indicador:
                    refs = refs_por_indicador[indicador["nome"]]
                    formula_str = indicador["formula"](refs) if refs else "0"
                    formula_aplicada = True
                elif indicador["nome"] in ["Nº ABANDONO(S)", "ABANDONO(S) (%)"]:
                    formula_str = indicador["formula"](dashboard_linha_atual, valor_col_letter)
                    formula_aplicada = True
                else:
                    formula_str = indicador["formula"]()
                    formula_aplicada = True
                if formula_aplicada:
                    cell_valor.value = formula_str
                    cell_valor.alignment = ALINHAMENTO_CENTRALIZADO
                    fmt = indicador.get("formato")
                    is_int = indicador["nome"] in indicadores_inteiros
                    if fmt:
                        cell_valor.number_format = fmt
                    elif is_int:
                        cell_valor.number_format = '0'
                    else:
                        cell_valor.number_format = '0.00'
            except Exception as e:
                logger.error(f"Erro formula SEC_GERAL '{indicador['nome']}': {e}")
                cell_valor.value = "#ERRO!"
        else:
            logger.warning(f"Formula não chamável SEC_GERAL '{indicador['nome']}'")
            cell_valor.value = "#N/A"
    linha_fim_dados_geral = dashboard_linha_atual
    for row in range(dashboard_linha_ref, linha_fim_dados_geral + 1):
        for col in range(dashboard_col_start_idx, dashboard_col_end_idx + 1):
            ws.cell(row=row, column=col).border = BORDER_THIN
    try:
        chart_status = BarChart()
        chart_status.type = "col"
        chart_status.style = 11
        title_font = Font(name='Calibri', size=14, bold=True)
        title_char_props = CharacterProperties(latin=title_font.name, sz=int(title_font.sz*100), b=title_font.bold)
        chart_status.title = RichTextProperties(p=[Paragraph(pPr=ParagraphProperties(defRPr=title_char_props), endParaRPr=title_char_props, r=[])], bodyPr='', lstStyle='', lvl1pPr='')
        chart_status.title.text = "Distribuição de Status Geral"
        chart_status.y_axis.title = 'Número de Alunos'
        chart_status.x_axis.title = 'Status'
        chart_status.height = 10
        chart_status.width = 15
        row_ativos = linha_inicio_dados_geral + 1
        row_transferidos = linha_inicio_dados_geral + 2
        row_desistentes = linha_inicio_dados_geral + 3
        if row_desistentes <= linha_fim_dados_geral:
            data = Reference(ws, min_col=dashboard_col_end_idx, min_row=row_ativos, max_row=row_desistentes)
            cats = Reference(ws, min_col=dashboard_col_start_idx, min_row=row_ativos, max_row=row_desistentes)
            chart_status.add_data(data, titles_from_data=False)
            chart_status.set_categories(cats)
            chart_status.legend = None
            series = chart_status.series[0]
            colors = ["9BBB59", "FFC000", "C0504D"]
            for idx, pt in enumerate(series.points):
                if idx < len(colors):
                    pt.graphicalProperties.solidFill = colors[idx]
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.numFmt = '0'
            chart_status.y_axis.majorGridlines = None
            chart_anchor = f"{get_column_letter(dashboard_col_start_idx + 3)}{dashboard_linha_ref}"
            ws.add_chart(chart_status, chart_anchor)
            logger.info(f"Gráfico status GERAL add {chart_anchor}")
        else:
            logger.warning(f"Dados insuficientes para gráfico status GERAL.")
    except Exception as chart_err:
        logger.error(f"Erro gráfico status GERAL: {chart_err}")

def criar_dashboard_sec_aprovacao(ws, turmas, linhas_inicio_tabelas_headers, start_col='R', start_row=4):
    """Cria o dashboard de aprovação e o GRÁFICO na aba SEC."""
    if not linhas_inicio_tabelas_headers or not turmas:
        logger.warning("Dados insuficientes para dash SEC APROVAÇÃO.")
        return
    dashboard_col_start_idx = column_index_from_string(start_col)
    num_bimestres = 4
    max_col_idx = dashboard_col_start_idx + num_bimestres
    dashboard_linha_ref = start_row
    cell_titulo = ws.cell(row=dashboard_linha_ref, column=dashboard_col_start_idx, value="TAXA MÉDIA APROVAÇÃO BIMESTRAL")
    cell_titulo.font = Font(bold=True, size=12)
    cell_titulo.alignment = ALINHAMENTO_CENTRALIZADO
    ws.merge_cells(start_row=dashboard_linha_ref, start_column=dashboard_col_start_idx, end_row=dashboard_linha_ref, end_column=max_col_idx)
    dashboard_linha_atual = dashboard_linha_ref + 1
    ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx, value="TURMA").font = Font(bold=True)
    ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx).alignment = ALINHAMENTO_CENTRALIZADO
    for i in range(num_bimestres):
        cell = ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx + 1 + i, value=f"B{i+1}")
        cell.font = Font(bold=True)
        cell.alignment = ALINHAMENTO_CENTRALIZADO
        cell.fill = FILL_BIMESTRES
    linha_inicio_dados_turmas = dashboard_linha_atual + 1
    wb = ws.parent
    for idx, turma in enumerate(turmas):
        dashboard_linha_atual += 1
        ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx, value=turma["nome_turma"]).alignment = Alignment(horizontal='left', vertical='center')
        linhas_por_bloco_turma_disc = 1 + 1 + MAX_ALUNOS_FORMATAR + 15
        linha_inicio_bloco_disc = 3 + (idx * linhas_por_bloco_turma_disc)
        linha_header_disc = linha_inicio_bloco_disc + 1
        linha_inicio_dash_disc = linha_header_disc + 1
        linha_ref_taxa_aprov = linha_inicio_dash_disc + 10
        col_map_disc_dash = {'O': 'N', 'P': 'O', 'Q': 'P', 'R': 'Q'}
        for disc_col_letter, aprov_col_letter in col_map_disc_dash.items():
            refs = [f"IFERROR('{d}'!{disc_col_letter}{linha_ref_taxa_aprov}, 0)" for d in DISCIPLINAS if d in wb.sheetnames]
            formula_str = f'=IFERROR(AVERAGE({",".join(refs)}),0)' if refs else '=0'
            aprov_col_idx = list(col_map_disc_dash.values()).index(aprov_col_letter) + dashboard_col_start_idx + 1
            cell = ws.cell(row=dashboard_linha_atual, column=aprov_col_idx)
            cell.value = formula_str
            cell.number_format = '0.00%'
            cell.alignment = ALINHAMENTO_CENTRALIZADO
    linha_fim_dados_turmas = dashboard_linha_atual
    for indicador_info in DASHBOARD_SEC_APROVACAO:
        dashboard_linha_atual += 1
        ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx, value=indicador_info["nome"]).font = Font(size=10, bold=True)
        ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx).alignment = ALINHAMENTO_CENTRALIZADO
        for i in range(num_bimestres):
            col_letter = get_column_letter(dashboard_col_start_idx + 1 + i)
            cell = ws.cell(row=dashboard_linha_atual, column=dashboard_col_start_idx + 1 + i)
            if indicador_info["nome"] == "TX APROVAÇÃO %":
                formula_str = f'=IFERROR(AVERAGE({col_letter}{linha_inicio_dados_turmas}:{col_letter}{linha_fim_dados_turmas}),0)' if linha_inicio_dados_turmas <= linha_fim_dados_turmas else '=0'
            elif indicador_info["nome"] == "TX REPROVAÇÃO %":
                formula_str = f'=IFERROR(1-{col_letter}{dashboard_linha_atual-1}, 0)'
            else:
                formula_str = '#N/A'
            cell.value = formula_str
            cell.font = Font(size=10, bold=True)
            cell.number_format = indicador_info["formato"]
            cell.alignment = ALINHAMENTO_CENTRALIZADO
    max_row_aprov = dashboard_linha_atual
    for row in range(dashboard_linha_ref, max_row_aprov + 1):
        for col in range(dashboard_col_start_idx, max_col_idx + 1):
            ws.cell(row=row, column=col).border = BORDER_THIN
    try:
        chart_aprov = BarChart()
        chart_aprov.type = "col"
        chart_aprov.style = 10
        title_font_aprov = Font(name='Calibri', size=14, bold=True)
        title_char_props_aprov = CharacterProperties(latin=title_font_aprov.name, sz=int(title_font_aprov.sz*100), b=title_font_aprov.bold)
        chart_aprov.title = RichTextProperties(p=[Paragraph(pPr=ParagraphProperties(defRPr=title_char_props_aprov), endParaRPr=title_char_props_aprov, r=[])], bodyPr='', lstStyle='', lvl1pPr='')
        chart_aprov.title.text = "TAXA MÉDIA DE APROVAÇÃO POR TURMA E BIMESTRE"
        chart_aprov.y_axis.title = "Taxa Média (%)"
        chart_aprov.x_axis.title = "Turma"
        chart_aprov.height = 14
        chart_aprov.width = 22
        min_col_data = dashboard_col_start_idx + 1
        max_col_data = max_col_idx
        min_row_data = linha_inicio_dados_turmas - 1
        max_row_data = linha_fim_dados_turmas
        if min_row_data < max_row_data:
            data = Reference(ws, min_col=min_col_data, min_row=min_row_data, max_col=max_col_data, max_row=max_row_data)
            cats = Reference(ws, min_col=dashboard_col_start_idx, min_row=linha_inicio_dados_turmas, max_row=max_row_data)
            chart_aprov.add_data(data, titles_from_data=True)
            chart_aprov.set_categories(cats)
            series_colors = ["4F81BD", "C0504D", "9BBB59", "8064A2"]
            for idx, series in enumerate(chart_aprov.series):
                if idx < len(series_colors):
                    series.graphicalProperties.solidFill = series_colors[idx]
            chart_aprov.legend.position = 'b'
            chart_aprov.y_axis.scaling.min = 0.0
            chart_aprov.y_axis.scaling.max = 1.0
            chart_aprov.y_axis.majorUnit = 0.2
            chart_aprov.y_axis.number_format = '0%'
            chart_aprov.y_axis.majorGridlines = ChartLines()
            chart_aprov.y_axis.majorGridlines.graphicalProperties.line.solidFill = "D9D9D9"
            chart_anchor = f"{get_column_letter(dashboard_col_start_idx + 6)}{dashboard_linha_ref}"
            ws.add_chart(chart_aprov, chart_anchor)
            logger.info(f"Gráfico aprovação add {chart_anchor}")
        else:
            logger.warning(f"Sem dados para gráfico aprovação ({ws.title}).")
    except Exception as chart_err:
        logger.error(f"Erro gráfico aprovação: {chart_err}")