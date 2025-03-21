import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

# Lista de disciplinas
DISCIPLINAS = ["BIO", "MAT", "FIS", "QUI", "GEO", "SOC", "HIST", "FIL", "ESP", "POR", "ART", "ADF", "ING"]

# Colunas da planilha
COLUNAS = [
    "Nº", "Nome do Aluno", "1º BIM", "2º BIM", "3º BIM", "4º BIM",
    "NF", "MG", "MF", "SITUAÇÃO DO ALUNO", "PF", "SF"
]

# Caminho da imagem
CAMINHO_IMAGEM = os.path.expanduser("~/teste-planilha-siage-interno/siage_interno.png")

# Função para configurar o comprimento das colunas
def configurar_largura_colunas(ws, colunas_largura):
    """
    Define a largura das colunas especificadas.
    :param ws: A worksheet (aba) onde as colunas serão configuradas.
    :param colunas_largura: Um dicionário onde a chave é o nome da coluna e o valor é a largura em cm.
    """
    for coluna_nome, largura_cm in colunas_largura.items():
        # Encontra o índice da coluna com base no nome
        coluna_idx = COLUNAS.index(coluna_nome) + 1  # +1 porque as colunas começam em 1 no Excel
        # Converte o índice para a letra da coluna (A, B, C, etc.)
        coluna_letra = get_column_letter(coluna_idx)
        # Converte a largura de cm para unidades do Excel (1 cm ≈ 3.78 unidades)
        largura_unidades = largura_cm * 3.78
        # Aplica a largura à coluna
        ws.column_dimensions[coluna_letra].width = largura_unidades

# Função para criar uma aba em branco (sem cabeçalho ou fórmulas)
def criar_aba_em_branco(wb, titulo, img):
    """
    Cria uma nova aba no Workbook com o título especificado,
    adiciona a imagem e o título, mas sem cabeçalho ou fórmulas.
    """
    ws = wb.create_sheet(title=titulo)
    
    # Mescla as células da primeira linha de A a J
    ws.merge_cells('A1:J1')
    
    # Ajusta a altura da linha mesclada para caber a imagem
    ws.row_dimensions[1].height = img.height * 0.75  # Ajusta a altura da linha (em pontos)
    
    # Adiciona a imagem na célula mesclada
    ws.add_image(img, 'A1')
    
    # Adiciona o texto "COMPOSITOR LUIS RAMALHO" na célula mesclada
    cell = ws['A1']
    cell.value = "COMPOSITOR LUIS RAMALHO"
    cell.font = Font(name='Arial', size=26, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    return ws

# Função para criar uma aba de disciplina (com cabeçalho e fórmulas)
def criar_aba_disciplina(wb, titulo, img):
    """
    Cria uma nova aba no Workbook com o título especificado,
    adiciona a imagem, o título, o cabeçalho e as fórmulas.
    """
    ws = wb.create_sheet(title=titulo)
    
    # Mescla as células da primeira linha de A a J
    ws.merge_cells('A1:J1')
    
    # Ajusta a altura da linha mesclada para caber a imagem
    ws.row_dimensions[1].height = img.height * 0.75  # Ajusta a altura da linha (em pontos)
    
    # Adiciona a imagem na célula mesclada
    ws.add_image(img, 'A1')
    
    # Adiciona o texto "COMPOSITOR LUIS RAMALHO" na célula mesclada
    cell = ws['A1']
    cell.value = "COMPOSITOR LUIS RAMALHO"
    cell.font = Font(name='Arial', size=26, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Desloca o cabeçalho para a linha 12
    for _ in range(10):  # Adiciona 10 linhas em branco
        ws.append([])
    
    # Adiciona o cabeçalho (nomes das colunas) na linha 12
    ws.append(COLUNAS)
    
    # Configura a largura das colunas específicas
    configurar_largura_colunas(ws, {
        "Nome do Aluno": 10,  # 10 cm
        "SITUAÇÃO DO ALUNO": 4.5  # 4.5 cm
    })
    
    # Adiciona os números de 1 a 35 na coluna "Nº" (a partir da linha 13)
    for i in range(1, 36):
        ws[f'A{i + 12}'] = i  # Linha 13 em diante
    
    # Adiciona fórmulas para calcular as médias e situações
    for row in range(13, 48):  # 35 alunos (linhas 13 a 47)
        # Fórmula para Nota Final (NF) - Média dos 4 bimestres
        ws[f'G{row}'] = f'=AVERAGE(C{row}:F{row})'
        
        # Fórmula para Média Geral (MG) - Média de todas as disciplinas
        ws[f'H{row}'] = f'=SUM(C{row}:F{row})/4'
        
        # Fórmula para Média Final (MF)
        ws[f'I{row}'] = f'=IF(H{row}<7, (0.6*H{row}) + (0.4*G{row}), "-")'
        
        # Fórmula para Situação do Aluno (SITUAÇÃO DO ALUNO)
        ws[f'J{row}'] = f'=IF(H{row}<2.5, "REPROVADO", IF(H{row}<7, "FINAL", "APROVADO"))'
        
        # Fórmula para Prova Final (PF)
        ws[f'K{row}'] = f'=IF(H{row}<7, (12.5 - (1.5*H{row})), "-")'
        
        # Fórmula para Situação Final (SF)
        ws[f'L{row}'] = f'=IF(G{row}>=K{row}, "AF", "-")'
    
    return ws

# Função principal
def criar_planilha():
    """
    Função principal para criar a planilha de notas.
    """
    # Cria um arquivo Excel usando openpyxl
    wb = Workbook()

    # Remove a sheet padrão criada automaticamente
    wb.remove(wb.active)

    # Verifica se a imagem existe
    if not os.path.exists(CAMINHO_IMAGEM):
        raise FileNotFoundError(f"A imagem não foi encontrada no caminho: {CAMINHO_IMAGEM}")

    # Carrega a imagem
    img = Image(CAMINHO_IMAGEM)

    # Reduz a imagem em 50%
    img.width = int(img.width * 0.5)  # Nova largura: 540 pixels
    img.height = int(img.height * 0.5)  # Nova altura: 106 pixels

    # Cria a aba "SEC" (Secretaria Escolar) em branco
    criar_aba_em_branco(wb, "SEC", img)

    # Cria uma sheet para cada disciplina (com cabeçalho e fórmulas)
    for disciplina in DISCIPLINAS:
        criar_aba_disciplina(wb, disciplina, img)

    # Cria as abas adicionais em branco
    abas_adicionais = ["INDIVIDUAL", "BOLETIM", "BOL", "RESULTADO", "FREQUÊNCIA"]
    for aba in abas_adicionais:
        criar_aba_em_branco(wb, aba, img)

    # Define o caminho onde o arquivo será salvo
    caminho_padrao = "/mnt/c/Users/lmbernardo/Downloads"  # Caminho no WSL2 para a pasta Downloads do Windows
    nome_arquivo = "planilha_notas_complexa.xlsx"
    caminho_completo = os.path.join(caminho_padrao, nome_arquivo)

    # Salva o arquivo Excel
    wb.save(caminho_completo)

    print(f"Arquivo Excel salvo em: {caminho_completo}")

# Executa a função principal
if __name__ == "__main__":
    criar_planilha()